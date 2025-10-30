import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

import os
from datetime import datetime
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pytz

from config import LEDGER_FILE
import config

LEDGER_COLUMNS = [
    "ID",
    "Type",
    "Date",
    "Timestamp (UTC)",
    "Party/Vendor",
    "Description",
    "Amount (₪)",
    "Payment Method",
    "Local Path",
    "Drive File ID",
    "Drive Folder ID",
    "Status",
    "Notes",
    "Created By",
]


class LedgerService:
    """Manages Excel ledger operations."""
    
    def __init__(self, ledger_file: str = LEDGER_FILE):
        """
        Initialize ledger service.
        
        Args:
            ledger_file: Path to Excel ledger file
        """
        self.ledger_file = ledger_file
        self._ensure_ledger_exists()
    
    def _ensure_ledger_exists(self):
        """Create ledger file if it doesn't exist."""
        if not os.path.exists(self.ledger_file):
            self._create_new_ledger()
    
    def _create_new_ledger(self):
        """Create a new ledger with headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        
        # Na Lav headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(LEDGER_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # ID
        ws.column_dimensions['B'].width = 12  # Type
        ws.column_dimensions['C'].width = 12  # Date
        ws.column_dimensions['D'].width = 20  # Timestamp
        ws.column_dimensions['E'].width = 25  # Party/Vendor
        ws.column_dimensions['F'].width = 40  # Description
        ws.column_dimensions['G'].width = 15  # Amount
        ws.column_dimensions['H'].width = 15  # Payment Method
        ws.column_dimensions['I'].width = 40  # Local Path
        ws.column_dimensions['J'].width = 45  # Drive File ID
        ws.column_dimensions['K'].width = 45  # Drive Folder ID
        ws.column_dimensions['L'].width = 12  # Status
        ws.column_dimensions['M'].width = 30  # Notes
        
        wb.save(self.ledger_file)
        print(f"✅ Created new ledger: {self.ledger_file}")
    
    def add_entry(
        self,
        entry_id: str,
        entry_type: str,
        amount: float,
        party: str,
        description: str,
        payment_method: str = "",
        local_path: str = "",
        drive_file_id: str = "",
        drive_folder_id: str = "",
        notes: str = "",
        created_by: str = "bot",
    ) -> bool:
        """
        Add a new entry to the ledger.
        
        Args:
            entry_id: Unique entry ID
            entry_type: Income, Expense, Pension, Study
            amount: Amount in shekels
            party: Client/vendor name
            description: Description
            payment_method: Payment method
            local_path: Local file path
            drive_file_id: Drive file ID
            drive_folder_id: Drive folder ID
            notes: Additional notes
            created_by: Creator identifier
        
        Returns:
            True if successful
        """
        # Check for duplicate ID
        if self._entry_exists(entry_id):
            print(f"⚠ Entry with ID '{entry_id}' already exists")
            return False
        
        wb = openpyxl.load_workbook(self.ledger_file)
        ws = wb.active
        
        # Prepare data row
        date = datetime.now().date()
        timestamp = datetime.now(pytz.UTC).isoformat()
        
        row_data = [
            entry_id,
            entry_type,
            date,
            timestamp,
            party,
            description,
            amount,
            payment_method,
            local_path,
            drive_file_id,
            drive_folder_id,
            "Active",
            notes,
            created_by,
        ]
        
        # Append row
        ws.append(row_data)
        
        # Style the amount column
        last_row = ws.max_row
        amount_cell = ws.cell(row=last_row, column=LEDGER_COLUMNS.index("Amount (₪)") + 1)
        amount_cell.number_format = '#,##0.00'
        
        wb.save(self.ledger_file)
        print(f"✅ Added entry: {entry_id}")
        return True
    
    def _entry_exists(self, entry_id: str) -> bool:
        """Check if entry ID already exists."""
        if not os.path.exists(self.ledger_file):
            return False
        
        wb = openpyxl.load_workbook(self.ledger_file)
        ws = wb.active
        
        id_column = LEDGER_COLUMNS.index("ID") + 1
        
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=id_column).value == entry_id:
                return True
        
        return False
    
    def get_last_entries(self, n: int = 5) -> List[Dict[str, Any]]:
        """
        Get last n entries from ledger.
        
        Args:
            n: Number of entries to retrieve
        
        Returns:
            List of entry dictionaries
        """
        if not os.path.exists(self.ledger_file):
            return []
        
        wb = openpyxl.load_workbook(self.ledger_file)
        ws = wb.active
        
        entries = []
        
        # Read from bottom up
        start_row = max(2, ws.max_row - n + 1)
        
        for row in range(start_row, ws.max_row + 1):
            entry = {}
            for col_num, column in enumerate(LEDGER_COLUMNS, 1):
                cell_value = ws.cell(row=row, column=col_num).value
                entry[column] = cell_value
            entries.append(entry)
        
        return entries
    
    def export_to_dict(self) -> List[Dict[str, Any]]:
        """Export entire ledger to list of dictionaries."""
        if not os.path.exists(self.ledger_file):
            return []
        
        wb = openpyxl.load_workbook(self.ledger_file)
        ws = wb.active
        
        entries = []
        
        for row in range(2, ws.max_row + 1):
            entry = {}
            for col_num, column in enumerate(LEDGER_COLUMNS, 1):
                cell_value = ws.cell(row=row, column=col_num).value
                entry[column] = cell_value
            entries.append(entry)
        
        return entries

    def add_entry_to_all_ledgers(
        self,
        entry_id: str,
        entry_type: str,
        amount: float,
        party: str,
        description: str,
        payment_method: str = "",
        local_path: str = "",
        drive_file_id: str = "",
        drive_folder_id: str = "",
        notes: str = "",
        created_by: str = "bot",
        year: int = None,
        month: int = None,
    ) -> bool:
        """
        Add entry to both monthly and yearly ledgers.
        
        Args:
            Same as add_entry, plus:
            year: Year for organizing (default: current year)
            month: Month for organizing (default: current month)
        
        Returns:
            True if successful
        """
        if year is None:
            year = config.get_current_year()
        if month is None:
            month = config.get_current_month()
        
        # Get paths for both ledgers
        monthly_ledger_path = config.get_monthly_ledger_path(year, month)
        yearly_ledger_path = config.get_yearly_ledger_path(year)
        
        success = True
        
        # Add to monthly ledger
        monthly_ledger = LedgerService(monthly_ledger_path)
        if not monthly_ledger.add_entry(
            entry_id=entry_id,
            entry_type=entry_type,
            amount=amount,
            party=party,
            description=description,
            payment_method=payment_method,
            local_path=local_path,
            drive_file_id=drive_file_id,
            drive_folder_id=drive_folder_id,
            notes=notes,
            created_by=created_by,
        ):
            print(f"⚠ Failed to add to monthly ledger: {monthly_ledger_path}")
            success = False
        
        # Add to yearly ledger
        yearly_ledger = LedgerService(yearly_ledger_path)
        if not yearly_ledger.add_entry(
            entry_id=entry_id,
            entry_type=entry_type,
            amount=amount,
            party=party,
            description=description,
            payment_method=payment_method,
            local_path=local_path,
            drive_file_id=drive_file_id,
            drive_folder_id=drive_folder_id,
            notes=notes,
            created_by=created_by,
        ):
            print(f"⚠ Failed to add to yearly ledger: {yearly_ledger_path}")
            success = False
        
        if success:
            print(f"✅ Added to monthly ledger: {monthly_ledger_path}")
            print(f"✅ Added to yearly ledger: {yearly_ledger_path}")
        
        return success


if __name__ == "__main__":
    # Test ledger service
    ledger = LedgerService()
    
    # Test adding entry
    test_id = "R-2025-0001"
    ledger.add_entry(
        entry_id=test_id,
        entry_type="Income",
        amount=2016.00,
        party="Algolight Ltd.",
        description="Kyūサービス業者サービス提供内容説明書および業務代理合同委任契約書",
        payment_method="Bank Transfer",
        notes="September 2025 services",
    )
    
    # Get last entries
    print("\nLast 5 entries:")
    entries = ledger.get_last_entries(5)
    for entry in entries:
        print(f"  {entry.get('ID')}: {entry.get('Type')} - {entry.get('Amount (₪)')}₪")

